"""Safety-requirements table — the Excel boundary of the chain.

One canonical column registry drives everything: the downloadable template
(Requirements sheet + Description sheet), deterministic row validation, the
conversion of FuSa-relevant rows into the SYS-REQ work product (same pattern
as the ReqIF import), and the Excel results workbook. The columns follow the
lifecycle: Item → HARA → Safety Goals → FSR → TSR → Design → Verification →
Safety Validation.
"""
from __future__ import annotations

import io
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

from . import ids

ASILS = ["QM", "A", "B", "C", "D"]
SRT = ["FSR", "TSR"]
YESNO = ["YES", "NO"]
APPROVAL = ["Draft", "In Review", "Approved", "Rejected"]


@dataclass(frozen=True)
class Column:
    name: str
    description: str
    stage: str                                   # lifecycle stage the column belongs to
    allowed: list[str] | None = None
    fusa_mandatory: bool = False                 # required when FuSa Relevant? = YES
    field: str | None = None                     # key in the SYS-REQ item grammar


COLUMNS = [
    Column("Requirement ID", "Unique id in PREFIX-nnn form (e.g. SR-001); becomes the item id in SYS-REQ. "
           "Leave blank — the template fills SR-nnn in as you type a row, and the import "
           "auto-assigns any id still missing.", "General"),
    Column("Requirement Text", "The requirement itself, one testable 'shall' statement.", "General", field="text"),
    Column("Requirement Type", "Functional / Non-functional / Interface / …", "General", field="type"),
    Column("FuSa Relevant?", "YES routes the row into the safety lifecycle; NO rows are kept but not imported.", "HARA", allowed=YESNO),
    Column("Safety Goal ID", "Safety goal this requirement traces to (from the HARA), e.g. SG-001.", "Safety Goals", fusa_mandatory=True, field="parent"),
    Column("Hazard ID", "Hazard from the HARA that the safety goal mitigates, e.g. HZ-001.", "HARA", field="hazard"),
    Column("ASIL", "QM, A, B, C or D — inherited from the safety goal.", "Safety Goals", allowed=ASILS, fusa_mandatory=True, field="asil"),
    Column("Safety Requirement Type", "FSR (functional, what) or TSR (technical, how).", "FSR / TSR", allowed=SRT, fusa_mandatory=True, field="srt"),
    Column("Failure Mode", "Failure mode the requirement addresses (feeds FMEA/FMEDA).", "Design", field="failure_mode"),
    Column("Safety Mechanism", "SM id from SM-CATALOG that detects/controls the failure, e.g. SM-002.", "Design", fusa_mandatory=True, field="sm"),
    Column("Safe State", "State the item transitions to on fault detection.", "FSR / TSR", fusa_mandatory=True, field="safe_state"),
    Column("Fault Detection Time", "Max time to detect the fault (with unit, e.g. 5 ms); FDT + FRT ≤ FTTI.", "FSR / TSR", field="fdt"),
    Column("Fault Reaction Time", "Max time from detection to safe state (with unit).", "FSR / TSR", field="frt"),
    Column("Diagnostic Coverage", "Claimed DC of the safety mechanism: 0–100% (or 0..1).", "Design", field="dc"),
    Column("Assumptions of Use", "SEooC assumptions the integrator must satisfy.", "Safety Validation", field="aou"),
    Column("Verification Method", "How compliance is verified (review, analysis, test, fault injection …).", "Verification", fusa_mandatory=True, field="verification"),
    Column("Verification Test ID", "Test case id proving verification, e.g. VT-001.", "Verification", fusa_mandatory=True, field="verification_test"),
    Column("Validation Method", "How the safety goal is validated at vehicle/item level.", "Safety Validation", fusa_mandatory=True, field="validation"),
    Column("Validation Test ID", "Test case id proving validation, e.g. VAL-001.", "Safety Validation", fusa_mandatory=True, field="validation_test"),
    Column("Architecture Element", "Element of the architecture the requirement is allocated to.", "Design", field="element"),
    Column("Software/Hardware Element", "SW, HW or SYS allocation of the element.", "Design", field="allocation"),
    Column("Traceability", "Upstream ids this requirement traces from (customer req, safety goal …).", "General", fusa_mandatory=True, field="trace"),
    Column("FuSa Approval Status", "Draft / In Review / Approved / Rejected.", "Safety Validation", allowed=APPROVAL, field="approval"),
]

LIFECYCLE = [
    ("Item", "input/item-definition.md", None),
    ("HARA", "HARA (hazards HZ-nnn, S/E/C, ASIL classification)", "HARA"),
    ("Safety Goals", "SADS (safety goals SG-nnn, parents in HARA)", "SADS"),
    ("FSR", "SYS-REQ (imported FSR rows), TSR agent input", "SYS-REQ"),
    ("TSR", "TSR work product", "TSR"),
    ("Design", "TSC, SM-CATALOG, HSR, HW-DESIGN", "HW-DESIGN"),
    ("Verification", "SCA-REPORT, SEC-SCAN, independent review", "SEC-SCAN"),
    ("Safety Validation", "VALIDATION-REPORT (this workbook)", None),
]

EXAMPLE_ROW = ["SR-001", "The BPSM shall signal an invalid pressure status within 10 ms of detecting a sense-IC fault.",
               "Functional", "YES", "SG-001", "HZ-001", "D", "TSR", "sense-IC stuck-at", "SM-002",
               "pressure invalid flag set on CAN", "5 ms", "5 ms", "99%",
               "host evaluates the status word every cycle", "fault injection test", "VT-001",
               "vehicle-level braking test", "VAL-001", "sense IC", "HW", "CR-002", "Draft"]


# ---- template ---------------------------------------------------------------

def build_template() -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Requirements"
    ws.append([c.name for c in COLUMNS])
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="DDE7F2")
    ws.append(EXAMPLE_ROW)
    # Requirement IDs create themselves: as soon as a row has content, column A shows the
    # next SR-nnn. The import assigns ids server-side too, so a blank column A never blocks.
    last = ws.cell(row=1, column=len(COLUMNS)).column_letter
    for row in range(3, 501):
        ws.cell(row=row, column=1).value = f'=IF(COUNTA(B{row}:{last}{row})=0,"","SR-"&TEXT(ROW()-1,"000"))'
    for idx, c in enumerate(COLUMNS, 1):
        ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = max(16, min(len(c.name) + 4, 34))
        if c.allowed:
            dv = DataValidation(type="list", formula1='"' + ",".join(c.allowed) + '"', allow_blank=True)
            ws.add_data_validation(dv)
            dv.add(f"{ws.cell(row=2, column=idx).column_letter}2:{ws.cell(row=2, column=idx).column_letter}500")

    _write_description_sheet(wb)
    return wb


def _write_description_sheet(wb: Workbook) -> None:
    ds = wb.create_sheet("Description")
    ds.append(["Column", "Description", "Allowed values", "Mandatory when FuSa Relevant = YES", "Lifecycle stage"])
    for cell in ds[1]:
        cell.font = Font(bold=True)
    for c in COLUMNS:
        ds.append([c.name, c.description, ", ".join(c.allowed) if c.allowed else "free text",
                   "yes" if c.fusa_mandatory else "no", c.stage])
    ds.column_dimensions["A"].width = 26
    ds.column_dimensions["B"].width = 80
    ds.column_dimensions["C"].width = 30
    ds.column_dimensions["E"].width = 20


def write_template(path: str | Path) -> Path:
    path = Path(path)
    build_template().save(path)
    return path


# ---- parse + validate -------------------------------------------------------

def parse(source) -> list[dict]:
    """Rows of the Requirements sheet as {column name: str}; `_row` keeps the sheet row number."""
    wb = load_workbook(source, read_only=True, data_only=True)
    ws = wb["Requirements"] if "Requirements" in wb.sheetnames else wb.active
    it = ws.iter_rows(values_only=True)
    headers = [str(h).strip() if h is not None else "" for h in next(it, [])]
    rows = []
    for n, values in enumerate(it, start=2):
        rec = {h: str(v).strip() if v is not None else "" for h, v in zip(headers, values)}
        if any(rec.get(c.name) for c in COLUMNS):
            rec["_row"] = n
            rows.append(rec)
    return rows


def normalise_ids(rows: list[dict], prefix: str | None = None) -> list[str]:
    """Canonicalise, assign and de-duplicate Requirement IDs in place; returns notes.

    The same id rules `ids.normalise_items` applies to authored work products, applied at the
    spreadsheet boundary: `sr-1` becomes `SR-001`, a blank cell is filled, a repeat is renumbered.
    An id is never a reason to reject an upload — the sheet is data entry, not a grammar exam.

    Deterministic: adopts the prefix of the first usable id (else SR) and numbers above the
    highest already held by that prefix, in sheet-row order, so nothing collides.
    """
    canon = [ids.canonical(r.get("Requirement ID", "")) for r in rows]
    if prefix is None:
        prefix = next((c.rsplit("-", 1)[0] for c in canon if c), "SR")
    n = max((int(c.rsplit("-", 1)[1]) for c in canon if c and c.rsplit("-", 1)[0] == prefix), default=0)

    notes, seen = [], set()
    for rec, new in zip(rows, canon):
        raw = rec.get("Requirement ID", "")
        why = None
        if new is None:
            why = "assigned" if not raw else f"'{raw}' is not an id"
        elif new in seen:
            why, new = f"'{new}' used twice", None
        if new is None:
            n += 1
            new = f"{prefix}-{n:03d}"
            notes.append(f"row {rec.get('_row', '?')}: {why} → {new}")
        elif new != raw:
            notes.append(f"row {rec.get('_row', '?')}: '{raw}' → {new} (spelling normalised)")
        seen.add(new)
        rec["Requirement ID"] = new
    return notes


def apply_ids(body: bytes, rows: list[dict]) -> bytes:
    """Write the (assigned) Requirement IDs back into the workbook so the saved
    input file carries them permanently — later parses see them as manual ids."""
    wb = load_workbook(io.BytesIO(body), data_only=True)
    ws = wb["Requirements"] if "Requirements" in wb.sheetnames else wb.active
    for rec in rows:
        ws.cell(row=rec["_row"], column=1).value = rec["Requirement ID"]
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _parse_dc(raw: str) -> float:
    v = float(raw.rstrip("%").strip())
    if raw.strip().endswith("%"):
        return v
    return v * 100 if v <= 1.0 else v


def validate_row(rec: dict) -> list[str]:
    rid = rec.get("Requirement ID", "")
    label = rid or f"row {rec.get('_row', '?')}"
    errors = []
    # Ids are not validated here: normalise_ids() has already made them canonical and unique.
    for c in COLUMNS:
        v = rec.get(c.name, "")
        if c.allowed and v and v not in c.allowed:
            errors.append(f"{label}: {c.name} '{v}' not in {'/'.join(c.allowed)}")
    if rec.get("FuSa Relevant?", "").upper() == "YES":
        for c in COLUMNS:
            if c.fusa_mandatory and not rec.get(c.name, ""):
                errors.append(f"{label}: {c.name} is mandatory for FuSa-relevant requirements")
    dc = rec.get("Diagnostic Coverage", "")
    if dc:
        try:
            if not 0.0 <= _parse_dc(dc) <= 100.0:
                raise ValueError
        except ValueError:
            errors.append(f"{label}: Diagnostic Coverage '{dc}' must be 0–100% (or 0..1)")
    return errors


def validate_rows(rows: list[dict]) -> list[str]:
    """Content errors only — an id problem is repaired by normalise_ids(), never reported here."""
    return [e for rec in rows for e in validate_row(rec)]


# ---- SYS-REQ work product ---------------------------------------------------

def fusa_rows(rows: list[dict]) -> list[dict]:
    return [r for r in rows if r.get("FuSa Relevant?", "").upper() == "YES"]


def to_work_product(rows: list[dict], work_product: str = "SYS-REQ", agent: str = "reqtable-import") -> str:
    out = [f"---\nid: {work_product}\ntitle: Imported safety requirements\nagent: {agent}\n"
           f"date: {datetime.now(timezone.utc).date().isoformat()}\nsource: safety-requirements.xlsx\nstatus: imported\n---\n",
           f"# {work_product}\n\nFuSa-relevant rows of input/safety-requirements.xlsx in the house grammar.\n\n## Items\n"]
    for rec in fusa_rows(rows):
        out.append(f"### {rec['Requirement ID']}")
        for c in COLUMNS:
            if c.field and rec.get(c.name):
                out.append(f"- {c.field}: {rec[c.name]}")
        out.append("")
    return "\n".join(out)


# ---- Excel results workbook -------------------------------------------------

def build_results(report, rows: list[dict]) -> Workbook:
    """report: fusa.report.ValidationReport; rows: parsed requirements (may be empty)."""
    from ..report import EVIDENCE_HEADER, _evidence_rows

    wb = Workbook()
    ok = Font(color="1D9A4E", bold=True)
    bad = Font(color="C0392B", bold=True)

    s = wb.active
    s.title = "Summary"
    s.append(["FuSa Validation Report"]); s["A1"].font = Font(bold=True, size=14)
    s.append(["Verdict", report.verdict]); s["B2"].font = ok if report.verdict == "RELEASABLE" else bad
    s.append(["Generated", report.generated]); s.append(["Model", report.model])
    s.append(["Dry run", str(report.dry_run).lower()]); s.append(["ASIL", report.asil])
    s.append([])
    s.append(["Release blockers" if report.reasons else "Release blockers", "none" if not report.reasons else ""])
    for r in report.reasons:
        s.append(["", r])
    s.append([])
    fr = fusa_rows(rows)
    s.append(["Requirements", len(rows)]); s.append(["FuSa relevant", len(fr)])
    for asil in ASILS:
        n = sum(1 for r in fr if r.get("ASIL") == asil)
        if n:
            s.append([f"  ASIL {asil}", n])
    s.append(["Approved", sum(1 for r in fr if r.get("FuSa Approval Status") == "Approved")])
    s.append([])
    s.append(["Lifecycle stage", "Evidence", "Status"])
    for cell in s[s.max_row]:
        cell.font = Font(bold=True)
    status_by_wp = {a.work_product: a.status for a in report.work_products}
    if fr:                              # SYS-REQ is imported, not authored by a planned agent
        status_by_wp.setdefault("SYS-REQ", "imported")
    for stage, evidence, wp in LIFECYCLE:
        st = "provided (input)" if wp is None and stage == "Item" else \
             report.verdict.lower() if wp is None else status_by_wp.get(wp, "—")
        s.append([stage, evidence, st])
    s.column_dimensions["A"].width = 22; s.column_dimensions["B"].width = 52; s.column_dimensions["C"].width = 18

    rq = wb.create_sheet("Requirements")
    rq.append([c.name for c in COLUMNS] + ["Validation Status", "Issues"])
    for cell in rq[1]:
        cell.font = Font(bold=True)
    for rec in rows:
        issues = validate_row(rec)
        row = [rec.get(c.name, "") for c in COLUMNS] + ["OK" if not issues else "ISSUES", "; ".join(issues)]
        rq.append(row)
        rq.cell(row=rq.max_row, column=len(COLUMNS) + 1).font = ok if not issues else bad

    ev = wb.create_sheet("Work products")
    ev.append(EVIDENCE_HEADER)
    for cell in ev[1]:
        cell.font = Font(bold=True)
    for row in _evidence_rows(report):
        ev.append(row)

    _write_description_sheet(wb)
    return wb


def results_bytes(report, rows: list[dict]) -> bytes:
    buf = io.BytesIO()
    build_results(report, rows).save(buf)
    return buf.getvalue()


def template_bytes() -> bytes:
    buf = io.BytesIO()
    build_template().save(buf)
    return buf.getvalue()


# ---- FMEDA input template ---------------------------------------------------

FMEDA_COLUMNS = ["element", "mode", "lam_fit", "category", "dc", "safety_mechanism"]
FMEDA_EXAMPLE = [
    ["sense IC", "stuck-at output", "8.0", "SR", "0.99", "SM-002"],
    ["sense IC", "drift", "6.0", "SR", "0.90", "SM-002"],
    ["supply", "overvoltage", "4.0", "SPF", "0.0", ""],
    ["Vref", "drift", "2.0", "MPF", "0.99", "SM-003"],
    ["housing", "mechanical wear", "1.0", "SAFE", "0.0", ""],
]


def fmeda_template_text() -> str:
    """The failure-mode CSV the metrics tool reads (SPFM / LFM / PMHF).

    category: SR (safety-related, split by dc) | MPF | SPF | RF | MPF_L | MPF_D | MPF_P | SAFE
    dc:       diagnostic coverage as a fraction, 0..1 (not a percentage)
    """
    from .metrics import CATEGORIES
    header = (f"# category: {' | '.join(sorted(CATEGORIES))}\n"
              "# dc: diagnostic coverage 0..1 (fraction, not %)\n"
              "# lam_fit: failure rate in FIT (failures per 1e9 h)\n")
    lines = [",".join(FMEDA_COLUMNS)] + [",".join(r) for r in FMEDA_EXAMPLE]
    return header + "\n".join(lines) + "\n"
