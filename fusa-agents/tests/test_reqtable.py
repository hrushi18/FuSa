import io

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook

from test_ui import wait_idle

XLSX = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def make_upload(rows):
    """Filled template: header row from the registry + given rows, as xlsx bytes."""
    from fusa.tools import reqtable
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Requirements"
    ws.append([c.name for c in reqtable.COLUMNS])
    for r in rows:
        ws.append([r.get(c.name, "") for c in reqtable.COLUMNS])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


GOOD_ROW = {
    "Requirement ID": "SR-001", "Requirement Text": "The BPSM shall signal invalid pressure within 10 ms.",
    "Requirement Type": "Functional", "FuSa Relevant?": "YES", "Safety Goal ID": "SG-001",
    "Hazard ID": "HZ-001", "ASIL": "D", "Safety Requirement Type": "TSR",
    "Failure Mode": "stuck-at", "Safety Mechanism": "SM-002", "Safe State": "pressure invalid flag set",
    "Fault Detection Time": "5 ms", "Fault Reaction Time": "5 ms", "Diagnostic Coverage": "99%",
    "Assumptions of Use": "host evaluates status word every cycle",
    "Verification Method": "fault injection test", "Verification Test ID": "VT-001",
    "Validation Method": "vehicle-level test", "Validation Test ID": "VAL-001",
    "Architecture Element": "sense IC", "Software/Hardware Element": "HW",
    "Traceability": "CR-002", "FuSa Approval Status": "Approved",
}

NON_FUSA_ROW = {"Requirement ID": "SR-002", "Requirement Text": "The housing shall be blue.",
                "Requirement Type": "Non-functional", "FuSa Relevant?": "NO"}


# ---- template ---------------------------------------------------------------

def test_template_has_requirements_and_description_sheets(tmp_path):
    from fusa.tools import reqtable
    path = tmp_path / "t.xlsx"
    reqtable.write_template(path)
    wb = load_workbook(path)
    assert "Requirements" in wb.sheetnames and "Description" in wb.sheetnames
    headers = [c.value for c in wb["Requirements"][1]]
    assert headers == [c.name for c in reqtable.COLUMNS]
    assert len(headers) == 23
    assert wb["Requirements"].max_row >= 2                    # worked example row
    desc_first_col = [r[0].value for r in wb["Description"].iter_rows(min_row=2)]
    assert "ASIL" in desc_first_col and "Safe State" in desc_first_col


# ---- parse + validate -------------------------------------------------------

def test_valid_rows_pass(tmp_path):
    from fusa.tools import reqtable
    rows = reqtable.parse(io.BytesIO(make_upload([GOOD_ROW, NON_FUSA_ROW])))
    assert reqtable.validate_rows(rows) == []


def test_duplicate_id_rejected():
    from fusa.tools import reqtable
    rows = reqtable.parse(io.BytesIO(make_upload([GOOD_ROW, GOOD_ROW])))
    assert any("duplicate" in e.lower() and "SR-001" in e for e in reqtable.validate_rows(rows))


def test_fusa_row_missing_asil_and_goal_rejected():
    from fusa.tools import reqtable
    bad = GOOD_ROW | {"Requirement ID": "SR-003", "ASIL": "", "Safety Goal ID": ""}
    errors = reqtable.validate_rows(reqtable.parse(io.BytesIO(make_upload([bad]))))
    assert any("SR-003" in e and "ASIL" in e for e in errors)
    assert any("SR-003" in e and "Safety Goal ID" in e for e in errors)


def test_bad_enum_values_rejected():
    from fusa.tools import reqtable
    bad = GOOD_ROW | {"Requirement ID": "SR-004", "ASIL": "E", "Safety Requirement Type": "XSR",
                      "FuSa Relevant?": "MAYBE"}
    errors = reqtable.validate_rows(reqtable.parse(io.BytesIO(make_upload([bad]))))
    joined = " | ".join(errors)
    assert "ASIL" in joined and "XSR" in joined and "MAYBE" in joined


def test_dc_out_of_range_rejected():
    from fusa.tools import reqtable
    bad = GOOD_ROW | {"Requirement ID": "SR-005", "Diagnostic Coverage": "150%"}
    assert any("SR-005" in e and "Diagnostic Coverage" in e
               for e in reqtable.validate_rows(reqtable.parse(io.BytesIO(make_upload([bad])))))


def test_non_fusa_row_needs_no_safety_fields():
    from fusa.tools import reqtable
    assert reqtable.validate_rows(reqtable.parse(io.BytesIO(make_upload([NON_FUSA_ROW])))) == []


def test_to_work_product_renders_fusa_rows_in_house_grammar():
    from fusa.tools import reqtable
    rows = reqtable.parse(io.BytesIO(make_upload([GOOD_ROW, NON_FUSA_ROW])))
    md = reqtable.to_work_product(rows)
    assert md.startswith("---")
    assert "### SR-001" in md
    assert "- parent: SG-001" in md
    assert "- asil: D" in md
    assert "- sm: SM-002" in md
    assert "SR-002" not in md                                  # non-FuSa rows stay out


# ---- server -----------------------------------------------------------------

@pytest.fixture
def client(workspace):
    from fusa.ui.server import create_app
    app = create_app(root=workspace, dry_run=True)
    with TestClient(app) as c:
        yield c


def test_template_download(client):
    r = client.get("/api/template/requirements")
    assert r.status_code == 200
    assert r.headers["content-type"].startswith(XLSX)
    wb = load_workbook(io.BytesIO(r.content))
    assert "Requirements" in wb.sheetnames and "Description" in wb.sheetnames


def test_upload_requirements_saves_sysreq_and_runs_chain(client, workspace):
    r = client.post("/api/input/requirements", content=make_upload([GOOD_ROW, NON_FUSA_ROW]),
                    headers={"content-type": XLSX})
    assert r.status_code == 202
    assert r.json()["rows"] == 2 and r.json()["fusa_relevant"] == 1
    assert (workspace / "input" / "safety-requirements.xlsx").exists()
    assert "### SR-001" in (workspace / "_generated" / "SYS-REQ" / "SYS-REQ.md").read_text()
    wait_idle(client)
    assert client.get("/api/status").json()["records"]["SADS"]["status"] == "reviewed"


def test_upload_requirements_bad_rows_rejected(client, workspace):
    target = workspace / "input" / "safety-requirements.xlsx"
    before = target.read_bytes() if target.exists() else None
    bad = GOOD_ROW | {"Requirement ID": "SR-009", "ASIL": "E"}
    r = client.post("/api/input/requirements", content=make_upload([bad]),
                    headers={"content-type": XLSX})
    assert r.status_code == 400
    assert any("SR-009" in d for d in r.json()["detail"])
    after = target.read_bytes() if target.exists() else None
    assert after == before                       # rejected upload leaves input/ untouched


def test_report_xlsx_download(client):
    client.post("/api/input/requirements", content=make_upload([GOOD_ROW]),
                headers={"content-type": XLSX})
    wait_idle(client)
    r = client.get("/report.xlsx")
    assert r.status_code == 200
    wb = load_workbook(io.BytesIO(r.content))
    for sheet in ("Summary", "Requirements", "Work products", "Description"):
        assert sheet in wb.sheetnames
    reqs = wb["Requirements"]
    headers = [c.value for c in reqs[1]]
    assert "Validation Status" in headers and "Issues" in headers
    summary_text = " ".join(str(c.value) for row in wb["Summary"].iter_rows() for c in row)
    assert "HARA" in summary_text and "Safety Validation" in summary_text   # lifecycle mapping
    fsr_status = next(r[2] for r in wb["Summary"].iter_rows(values_only=True) if r and r[0] == "FSR")
    assert fsr_status != "—"                   # FSR stage reflects the uploaded SYS-REQ


def test_dashboard_offers_template_and_excel_report(client):
    html = client.get("/").text
    assert "/api/template/requirements" in html
    assert "/report.xlsx" in html


# ---- CLI --------------------------------------------------------------------

def test_cli_template_writes_file(workspace, tmp_path):
    import fusa.cli
    out = tmp_path / "template.xlsx"
    assert fusa.cli.main(["template", "--out", str(out)]) == 0
    assert load_workbook(out)["Requirements"] is not None
